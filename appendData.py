from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
import math
from openpyxl.styles import PatternFill
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart import ScatterChart, Reference
from openpyxl import Workbook, load_workbook
from genData import avg_generator





def subscr(num):
    return str(num).translate(str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉"))


def get_min_max(min, max):
    print(min, max)
    if min > max:
        min, max = max, min
    if abs(max - min) < 0.1:
        dec = 1000
    elif abs(max - min) < .8:
        dec = 100
    elif abs(max - min) < 2:
        dec = 10
    else:
        dec = 1
    print(f'min: {min}; max: {max}')
    min = math.floor(min * dec) / dec
    max = math.ceil(max * dec) / dec
    return min, max



def append_data(data, num_of_expts, indvar):
    wb_dest = load_workbook(f'templates/Avg_Template.xlsx')
    ws_dest = wb_dest['temp']
    exptnum1 = 0
    #if len(indvar) == 0, then there all the variables are the same in each experiment.
    if len(indvar) == 0:
        avg_template = True
    else:
        avg_template = False
        del ws_dest.tables['stddevtab']
        del ws_dest.tables['stats_table']
        for row in ws_dest[f'E26:L28']:
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fgColor="E7E6E6", fill_type="solid")


    if avg_template == True:
        stats_data = avg_generator(data, num_of_expts)
        len_stats = len(stats_data[1])
        print('len_stats: ',len_stats)
        maxrow = 29 + len_stats-1
        tabref = f"E28:L{maxrow}"
        ws_dest.tables[f"stats_table"].ref = tabref
        it1 = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxrow, min_col=5, max_col=5):
            for cell in row:
                cell.value = stats_data[0][it1]
                it1 += 1
        it1 = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxrow, min_col=6, max_col=6):
            for cell in row:
                cell.value = stats_data[1][it1]
                it1 += 1
        it1 = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxrow, min_col=7, max_col=7):
            for cell in row:
                cell.value = stats_data[2][it1]
                it1 += 1
        it1 = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxrow, min_col=8, max_col=8):
            for cell in row:
                cell.value = stats_data[3][it1]
                it1 += 1
        it1 = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxrow, min_col=9, max_col=9):
            for cell in row:
                cell.value = stats_data[4][it1]
                it1 += 1
        it1 = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxrow, min_col=10, max_col=10):
            for cell in row:
                cell.value = stats_data[5][it1]
                it1 += 1
        it1 = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxrow, min_col=11, max_col=11):
            for cell in row:
                cell.value = stats_data[6][it1]
                it1 += 1
        it1 = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxrow, min_col=12, max_col=12):
            for cell in row:
                cell.value = stats_data[7][it1]
                it1 += 1



    # data: [8:rpm],[9:electrode],[10: molarity],[11: temp],[12:electrolyte]
    print('success 2')
    position_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                     'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ',
                     'AK', 'AL', 'AM', 'AN', 'AO', 'AP']
    while exptnum1 < num_of_expts:
        temp = data[exptnum1][4][11]
        electrolyte = data[exptnum1][4][12]
        rsquared = data[exptnum1][4][13]
        molarity = data[exptnum1][4][10]
        electrode = data[exptnum1][4][9]
        rpm = data[exptnum1][4][8]
        curs = data[exptnum1][0]
        pots = data[exptnum1][1]
        curdens = data[exptnum1][2]
        logs = data[exptnum1][3]
        print(f'logs: {logs}')
        OERHER = data[0][4][0]
        exptname_1 = data[exptnum1][4][1]
        Tslope = data[exptnum1][4][2]
        range = data[exptnum1][4][3]
        ind1 = data[exptnum1][4][4]
        ind2 = data[exptnum1][4][5]
        log_low = min(logs)
        log_high = max(logs)
        pot_low = min(pots)
        pot_high = max(pots)
        curdens_low = min(curdens)
        curdens_high = max(curdens)
        set_len = len(pots)

        for char in electrolyte:
            if char == '_':
                ind = electrolyte.index(char)
                if ind + 1 != -1:
                    electrolyte = f'{electrolyte[:ind]}{subscr(electrolyte[ind + 1])}{electrolyte[ind + 2:]}'
                else:
                    electrolyte = f'{electrolyte[:ind]}{subscr(electrolyte[ind + 1])}'
        if avg_template == False:
            exptname_dict = {'°C': f'{temp}°C',
                             'M': f', {molarity}M',
                             #to add electrolyte to experiment name, remove the following line and paste {electrolyte} into the above string.
                             'Electrolyte': f'{electrolyte}',
                             'Electrode': f' {electrode}',
                             'RPM': f', {rpm} RPM'}
            exptname = ''.join(exptname_dict.get(key, '') for key in indvar)
            if exptname.startswith(','):
                exptname = exptname[2:]
        else:
            exptname = f'Expt. {exptnum1 + 1}'


        if temp == 'rt':
            temp = "RT"

        if exptnum1 == 0:
            exptnum = 14
        else:
            exptnum = (exptnum1 * 5) + 14
        num = 0

        maxro = 28 + set_len
        for row in ws_dest.iter_cols(min_row=29, max_row=maxro, min_col=exptnum, max_col=exptnum):
            for cell in row:
                cell.value = pots[num]
                num += 1
        num = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxro, min_col=exptnum + 1, max_col=exptnum + 1):
            for cell in row:
                cell.value = curs[num]
                num += 1
        num = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxro, min_col=exptnum + 2, max_col=exptnum + 2):
            for cell in row:
                cell.value = logs[num]
                if cell.value in logs[ind1:ind2 + 1]:
                    cell.fill = PatternFill(fgColor="ffd966", fill_type="solid")
                num += 1
        num = 0
        for row in ws_dest.iter_cols(min_row=29, max_row=maxro, min_col=exptnum + 3, max_col=exptnum + 3):
            for cell in row:
                cell.value = curdens[num]
                num += 1
        ws_dest[f'{position_list[exptnum]}25'].value = exptname_1
        ws_dest[f'{position_list[exptnum]}26'].value = temp
        ws_dest[f'{position_list[exptnum]}27'].value = rpm
        if avg_template == True:
            ws_dest['E27'].value = temp
            ws_dest['F27'].value = rpm
        ws_dest[f'{position_list[exptnum + 2]}26'].value = f'auto: {Tslope}'
        ws_dest[f'{position_list[exptnum + 2]}27'].value = f'{range:.2f}'
        tabref = f"{position_list[exptnum - 1]}28:{position_list[exptnum + 2]}"
        tabref = tabref + str(maxro)
        ws_dest.tables[f"expt_table{exptnum1}"].ref = tabref
        wb_dest.save('excelFiles/file_temp.xlsx')
        print(f'electrolyte: {electrolyte}')

        potsminmax = get_min_max(pot_low, pot_high)
        logsminmax = get_min_max(log_low, log_high)
        cursminmax = get_min_max(curdens_low, curdens_high)

        if exptnum1 == 0:
            line_color = 'C81C0B'
            line_color2 = 'EB5D50'
            dash_style = 'dot'
        elif exptnum1 == 1:
            line_color = '34AD23'
            line_color2 = '9BE77D'
            dash_style = 'lgDash'
        elif exptnum1 == 2:
            line_color = '0CBBC4'
            line_color2 = '6CE2DD'
            dash_style = 'dash'
        elif exptnum1 == 3:
            line_color = '833CDD'
            line_color2 = 'B088F3'
            dash_style = 'sysDash'
        elif exptnum1 == 4:
            line_color = 'CAC10C'
            line_color2 = 'EFE27C'
            dash_style = 'solid'
        elif exptnum1 == 5:
            line_color = 'E559B0'
            line_color2 = 'F6A8C9'
            dash_style = 'lgDashDotDot'

        if temp != 'rt' or temp != 'RT':
            temp = str(temp) + '°C'
        elif temp == 'rt':
            temp = 'RT'

        chart_title = f'{temp}, {rpm} RPM, {molarity}M {electrolyte}, {electrode} Electrode, {OERHER}'
        if avg_template == True:
            chart_title4 = f'Avg. {temp}, {rpm} RPM, {molarity}M {electrolyte}, {electrode} Electrode, {OERHER}'
        elif avg_template == False:
            chart_title4 = ''
            if '°C' not in indvar:
                chart_title4 = f'{temp}'
            if 'M' not in indvar or 'Electrolyte' not in indvar:
                chart_title4 = chart_title4 + f', {molarity}M {electrolyte}'
            if 'Electrode' not in indvar:
                chart_title4 = chart_title4 + f', {electrode} Electrode'
            if 'RPM' not in indvar:
                chart_title4 = chart_title4 + f', {rpm} RPM'


        num = 0
        if exptnum1 == 0:
            tafval = 5
        else:
            tafval = 5 + (exptnum1 * 2)

        if exptnum1 == 0:
            exptnum = 14
        else:
            exptnum = (exptnum1 * 5) + 14

        if exptnum1 == 0 and avg_template == True:
            len_stats = len(stats_data[1])
            maxrow = 29 + len_stats - 1
            for chart in ws_dest._charts:
                if chart._id == 18:
                    chart.title = f'{chart_title} LSV'
                    chart.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(
                        latin=Font(typeface='Calibre'), sz=1400, b=True))
                    avg_y = Reference(ws_dest, min_col=6, max_col=6, min_row=29, max_row=maxrow)
                    avg_x = Reference(ws_dest, min_col=5, max_col=5, min_row=29, max_row=maxrow)
                    avg_series = Series(avg_y, avg_x, title_from_data=False, title=f'avg.')
                    avg_series.graphicalProperties.line = LineProperties(prstDash='solid', w=16000, solidFill='C0C0C0')
                    chart.append(avg_series)

                    minerror_y = Reference(ws_dest, min_col=12, max_col=12, min_row=29, max_row=maxrow)
                    pos_error = Series(minerror_y, avg_x, title_from_data=False, title=f'-std. error')
                    pos_error.graphicalProperties.line = LineProperties(prstDash='solid', w=9000, solidFill='C0C0C0')
                    chart.append(pos_error)

                    maxerror_y = Reference(ws_dest, min_col=11, max_col=11, min_row=29, max_row=maxrow)
                    neg_error = Series(maxerror_y, avg_x, title_from_data=False, title=f'+std. error')
                    neg_error.graphicalProperties.line = LineProperties(prstDash='solid', w=9000, solidFill='C0C0C0')
                    chart.append(neg_error)
                    wb_dest.save('excelFiles/file_temp.xlsx')


        for chart in ws_dest._charts:
            if avg_template == True:
                index_num = exptnum1 + 3
            else:
                index_num = exptnum1

            if chart._id == 17:
                cell_low_ind = ind1 + 29
                cell_high_ind = ind2 + 29
                all_y = Reference(ws_dest, min_col=exptnum + 2, max_col=exptnum + 2, min_row=29, max_row=set_len + 29)
                all_x = Reference(ws_dest, min_col=exptnum + 0, max_col=exptnum + 0, min_row=29, max_row=set_len + 29)
                tafel_y = Reference(ws_dest, min_col=exptnum + 2, max_col=exptnum + 2, min_row=cell_low_ind, max_row=cell_high_ind)
                tafel_x = Reference(ws_dest, exptnum + 0, max_col=exptnum + 0, min_row=cell_low_ind, max_row=cell_high_ind)
                all_series = Series(all_y, all_x, title_from_data=False, title=f'delete')
                all_series.graphicalProperties.line = LineProperties(prstDash='solid', w=9000, solidFill='00000')
                tafel_series = Series(tafel_y, tafel_x, title_from_data=False, title=f'{exptname}, T slope {Tslope}, R²: {round(rsquared, 4)}')
                tafel_series.graphicalProperties.line = LineProperties(prstDash='solid', w=12000, solidFill=line_color)
                tafel_series.trendline = Trendline(dispRSqr=False, dispEq=False, forward=1, backward=1,
                                                   spPr=GraphicalProperties(ln=LineProperties(prstDash='lgDash', solidFill=line_color, w=7000)))

                chart.append(all_series)
                chart.append(tafel_series)
                chart.x_axis.scaling.min = potsminmax[0]
                chart.y_axis.scaling.min = logsminmax[0]
                chart.x_axis.scaling.max = potsminmax[1]
                chart.y_axis.scaling.max = logsminmax[1]
                wb_dest.save('excelFiles/file_temp.xlsx')

            elif chart._id == tafval:
                print('these next steps are sometimes slow')
                print(f'ind1: {ind1}')
                print(f'ind2: {ind2}')
                cell_low_ind = ind1 + 29
                cell_high_ind = ind2 + 29
                all_y = Reference(ws_dest, min_col=exptnum + 2, max_col=exptnum + 2, min_row=29, max_row=set_len + 29)
                all_x = Reference(ws_dest, min_col=exptnum + 0, max_col=exptnum + 0, min_row=29, max_row=set_len + 29)
                tafel_y = Reference(ws_dest, min_col=exptnum + 2, max_col=exptnum + 2, min_row=cell_low_ind,
                                    max_row=cell_high_ind)
                tafel_x = Reference(ws_dest, exptnum + 0, max_col=exptnum + 0, min_row=cell_low_ind,
                                    max_row=cell_high_ind)
                all_series = Series(all_y, all_x, title_from_data=False, title="total_series")
                tafel_series = Series(tafel_y, tafel_x, title_from_data=False, title="tafel_series")
                wb_dest.save('excelFiles/file_temp.xlsx')
                chart.append(all_series)
                chart.series[0] = all_series
                chart.series[0].graphicalProperties.line = LineProperties(w=15875, solidFill='000000')
                chart.series[1] = tafel_series
                chart.series[1].graphicalProperties.line = LineProperties(w=25400, solidFill='00A9DA')
                chart.series[1].trendline = Trendline(dispRSqr=True, dispEq=True, forward=1, backward=1,
                                                      spPr=GraphicalProperties(ln=LineProperties(prstDash='lgDash', solidFill='9C9C9C', w=12500)))
                chart.x_axis.scaling.min = potsminmax[0]
                chart.y_axis.scaling.min = logsminmax[0]
                chart.x_axis.scaling.max = potsminmax[1]
                chart.y_axis.scaling.max = logsminmax[1]
                chart.title = f'{chart_title} Tafel Plot'
                chart.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(
                    latin=Font(typeface='Calibre'), sz=1400, b=True))
                wb_dest.save('excelFiles/file_temp.xlsx')


            elif chart._id == tafval + 1:
                all_y = Reference(ws_dest, min_col=exptnum + 3, max_col=exptnum + 3, min_row=29, max_row=set_len + 29)
                all_x = Reference(ws_dest, min_col=exptnum, max_col=exptnum, min_row=29, max_row=set_len + 29)
                all_series = Series(all_y, all_x, title_from_data=False, title=exptname)
                chart.series[0] = all_series
                chart.series[0].graphicalProperties.line = LineProperties(w=15875, solidFill=line_color)
                chart.x_axis.scaling.min = potsminmax[0]
                chart.x_axis.scaling.max = potsminmax[1]
                chart.y_axis.scaling.min = cursminmax[0]
                chart.y_axis.scaling.max = cursminmax[1]
                chart.title = f'{chart_title} LSV'
                chart.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(
                    latin=Font(typeface='Calibre'), sz=1400, b=True))
                wb_dest.save('excelFiles/file_temp.xlsx')


            elif chart._id == 4:
                all_y = Reference(ws_dest, min_col=exptnum + 3, max_col=exptnum + 3, min_row=29, max_row=set_len + 29)
                all_x = Reference(ws_dest, min_col=exptnum, max_col=exptnum, min_row=29, max_row=set_len + 29)
                all_series = Series(all_y, all_x, title_from_data=False, title=f'{exptname}')
                chart.append(all_series)
                # This uses the same potential as the last experimental group and is not ideal
                chart.x_axis.scaling.min = potsminmax[0]
                chart.x_axis.scaling.max = potsminmax[1]
                chart.y_axis.scaling.min = cursminmax[0]
                chart.y_axis.scaling.max = cursminmax[1]
                chart.series[exptnum1] = all_series
                chart.series[exptnum1].graphicalProperties.line = LineProperties(prstDash=dash_style, w=15875, solidFill=line_color)
                wb_dest.save('excelFiles/file_temp.xlsx')

            elif chart._id == 3:
                cell_low_ind = ind1 + 29
                cell_high_ind = ind2 + 29
                all_y = Reference(ws_dest, min_col=exptnum + 2, max_col=exptnum + 2, min_row=29, max_row=set_len + 29)
                all_x = Reference(ws_dest, min_col=exptnum + 0, max_col=exptnum + 0, min_row=29, max_row=set_len + 29)
                tafel_y = Reference(ws_dest, min_col=exptnum + 2, max_col=exptnum + 2, min_row=cell_low_ind,
                                    max_row=cell_high_ind)
                tafel_x = Reference(ws_dest, exptnum + 0, max_col=exptnum + 0, min_row=cell_low_ind,
                                    max_row=cell_high_ind)
                all_series = Series(all_y, all_x, title_from_data=False, title=f'{exptname}, T slope {Tslope}, R²: {round(rsquared, 4)}')
                all_series.graphicalProperties.line = LineProperties(prstDash=dash_style, w=17000, solidFill=line_color)
                tafel_series = Series(tafel_y, tafel_x, title_from_data=False, title=f'delete')
                tafel_series.graphicalProperties.line = LineProperties(w=17000, solidFill=line_color2)

                chart.append(all_series)
                chart.append(tafel_series)
                # This uses the same potential as the last experimental group and is not ideal
                chart.x_axis.scaling.min = potsminmax[0]
                chart.y_axis.scaling.min = logsminmax[0]
                chart.x_axis.scaling.max = potsminmax[1]
                chart.y_axis.scaling.max = logsminmax[1]
                wb_dest.save('excelFiles/file_temp.xlsx')

            elif chart._id == 18 and avg_template == True:
                all_y = Reference(ws_dest, min_col=exptnum + 3, max_col=exptnum + 3, min_row=29, max_row=set_len + 29)
                all_x = Reference(ws_dest, min_col=exptnum, max_col=exptnum, min_row=29, max_row=set_len + 29)
                all_series = Series(all_y, all_x, title_from_data=False, title=f'{exptname}')
                chart.append(all_series)
                chart.x_axis.scaling.min = potsminmax[0]
                chart.x_axis.scaling.max = potsminmax[1]
                chart.y_axis.scaling.min = cursminmax[0]
                chart.y_axis.scaling.max = cursminmax[1]
                chart.series[index_num] = all_series
                chart.series[index_num].graphicalProperties.line = LineProperties(prstDash=dash_style, w=15875, solidFill=line_color)
                wb_dest.save('excelFiles/file_temp.xlsx')


        exptnum1 +=1

    for chart in ws_dest._charts:

        if chart._id == 4:
            chart.title = f'{chart_title} LSV'
            chart.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(
                latin=Font(typeface='Calibre'), sz=1400, b=True))
            wb_dest.save('excelFiles/file_temp.xlsx')

        elif chart._id == 3:
            chart.title = f'{chart_title} Tafel Plot'
            chart.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(
                latin=Font(typeface='Calibre'), sz=1400, b=True))
            wb_dest.save('excelFiles/file_temp.xlsx')

        elif chart._id == 17:
            chart.title = f'{chart_title} Tafel Plot'
            chart.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(
                latin=Font(typeface='Calibre'), sz=1400, b=True))
            wb_dest.save('excelFiles/file_temp.xlsx')




    number = num_of_expts
    print(f'num_of_expts {num_of_expts}')

    while number <= 5:
        print(f"expt_table{number}")
        del ws_dest.tables[f"expt_table{number}"]
        number += 1


    number = (num_of_expts * 2) + 2
    print(f'number {number}')
    del ws_dest._charts[number:14]
    ##commented this out cause it was slow
    ##print(f'this next step is sometimes slow')
    ##if num_of_expts < 6:
        ##ws_dest.delete_cols(13+ (5*num_of_expts), (5 * (6 - (num_of_expts))))
    ##print(f'ws_dest.delete_cols({13 + (5 * num_of_expts)}, {(5 * (6 - (num_of_expts)))})')
    for row in ws_dest[f'B{2+num_of_expts}:C7']:
        for cell in row:
            cell.value = None
    wb_dest.save('excelFiles/file_temp.xlsx')
    print('done2')



